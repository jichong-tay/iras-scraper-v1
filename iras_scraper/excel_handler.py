"""Excel file handler for reading UENs and writing results."""

import pandas as pd
from typing import List, Dict, Any
import logging
from pathlib import Path


class ExcelHandler:
    """Handles Excel file operations for UEN input and result output."""
    
    def __init__(self, input_file: str, output_file: str):
        """Initialize Excel handler.
        
        Args:
            input_file: Path to input Excel file containing UENs
            output_file: Path to output Excel file for results
        """
        self.input_file = Path(input_file)
        self.output_file = Path(output_file)
        self.logger = logging.getLogger(__name__)
    
    def read_uens(self, column_name: str = "UEN") -> List[str]:
        """Read UENs from Excel file.
        
        Args:
            column_name: Name of the column containing UENs
            
        Returns:
            List of UEN strings
            
        Raises:
            FileNotFoundError: If input file doesn't exist
            KeyError: If UEN column doesn't exist
        """
        try:
            if not self.input_file.exists():
                raise FileNotFoundError(f"Input file not found: {self.input_file}")
            
            df = pd.read_excel(self.input_file)
            
            if column_name not in df.columns:
                available_columns = ", ".join(df.columns.tolist())
                raise KeyError(f"Column '{column_name}' not found. Available columns: {available_columns}")
            
            uens = df[column_name].dropna().astype(str).tolist()
            self.logger.info(f"Successfully read {len(uens)} UENs from {self.input_file}")
            
            return uens
            
        except Exception as e:
            self.logger.error(f"Error reading UENs from Excel: {str(e)}")
            raise
    
    def write_results(self, results: List[Dict[str, Any]], overwrite: bool = True):
        """Write scraping results to Excel file with expanded data columns.
        
        Args:
            results: List of dictionaries containing scraping results
            overwrite: Whether to overwrite existing output file
        """
        try:
            # Create output directory if it doesn't exist
            self.output_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Convert results to DataFrame and expand data columns
            df = self._expand_data_columns(results)
            
            # Rename columns to be more user-friendly
            df_formatted = df.copy()
            column_mapping = {
                'uen': 'UEN',
                'success': 'Success',
                'gst_registration_status': 'GST Registration Status', 
                'business_status': 'Business Status',
                'company_name': 'Company Name',
                'entity_type': 'Entity Type',
                'registration_date': 'Registration Date',
                'extraction_status': 'Data Extraction Status',
                'raw_search_results': 'Raw Search Results',
                'page_title': 'Page Title',
                'page_url': 'Page URL',
                'extraction_error': 'Extraction Error',
                'error': 'Error Message',
                'timestamp': 'Timestamp'
            }
            
            # Apply column renaming for existing columns
            df_formatted = df_formatted.rename(columns={k: v for k, v in column_mapping.items() if k in df_formatted.columns})
            
            # Write to Excel with enhanced formatting
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                df_formatted.to_excel(writer, sheet_name='IRAS_Search_Results', index=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['IRAS_Search_Results']
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for col_num, column_title in enumerate(df_formatted.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in df_formatted:
                    column_width = max(df_formatted[column].astype(str).map(len).max(), len(column)) + 4
                    col_letter = worksheet.cell(row=1, column=df_formatted.columns.get_loc(column) + 1).column_letter
                    worksheet.column_dimensions[col_letter].width = min(column_width, 60)
                
                # Set specific widths for certain columns
                specific_widths = {
                    'UEN': 15,
                    'Success': 10, 
                    'GST Registration Status': 25,
                    'Business Status': 20,
                    'Company Name': 40,
                    'Entity Type': 20,
                    'Registration Date': 18,
                    'Data Extraction Status': 20,
                    'Raw Search Results': 80,
                    'Error Message': 30,
                    'Timestamp': 20
                }
                
                for col_name, width in specific_widths.items():
                    if col_name in df_formatted.columns:
                        col_letter = worksheet.cell(row=1, column=df_formatted.columns.get_loc(col_name) + 1).column_letter
                        worksheet.column_dimensions[col_letter].width = width
            
            self.logger.info(f"Successfully wrote {len(results)} results to {self.output_file}")
            
        except Exception as e:
            self.logger.error(f"Error writing results to Excel: {str(e)}")
            raise
    
    def _expand_data_columns(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """Expand the nested 'data' column into separate columns for better readability.
        
        Args:
            results: List of dictionaries containing scraping results
            
        Returns:
            DataFrame with expanded columns
        """
        expanded_results = []
        
        for result in results:
            # Start with the base result
            expanded_result = {
                'uen': result.get('uen', ''),
                'success': result.get('success', False),
                'error': result.get('error', None),
                'timestamp': result.get('timestamp', '')
            }
            
            # Extract and expand the data dictionary if it exists
            data_dict = result.get('data', {})
            if isinstance(data_dict, dict):
                # Add key data fields as separate columns with user-friendly names
                expanded_result.update({
                    'gst_registration_status': data_dict.get('gst_registration', ''),
                    'business_status': data_dict.get('status', ''),
                    'company_name': data_dict.get('company_name', ''),
                    'entity_type': data_dict.get('entity_type', ''),
                    'registration_date': data_dict.get('registration_date', ''),
                    'extraction_status': data_dict.get('extraction_status', ''),
                    'raw_search_results': data_dict.get('raw_results', ''),
                    'page_title': data_dict.get('page_title', ''),
                    'page_url': data_dict.get('page_url', ''),
                    'extraction_error': data_dict.get('extraction_error', ''),
                })
                
                # Add any additional fields that might be present
                for key, value in data_dict.items():
                    if key not in ['gst_registration', 'status', 'company_name', 'entity_type', 
                                 'registration_date', 'extraction_status', 'raw_results', 
                                 'page_title', 'page_url', 'extraction_error']:
                        expanded_result[f'data_{key}'] = value
            else:
                # If data is not a dict, store it as-is
                expanded_result['data_raw'] = str(data_dict) if data_dict else ''
            
            expanded_results.append(expanded_result)
        
        # Create DataFrame with consistent column order
        column_order = [
            'uen', 'success', 'gst_registration_status', 'business_status', 
            'company_name', 'entity_type', 'registration_date', 'extraction_status',
            'raw_search_results', 'page_title', 'page_url', 'extraction_error', 
            'error', 'timestamp'
        ]
        
        df = pd.DataFrame(expanded_results)
        
        # Reorder columns (only include existing ones)
        existing_columns = [col for col in column_order if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in existing_columns]
        final_column_order = existing_columns + remaining_columns
        
        return df[final_column_order]
    
    def append_results(self, new_results: List[Dict[str, Any]]):
        """Append new results to existing Excel file.
        
        Args:
            new_results: List of new results to append
        """
        try:
            existing_results = []
            
            # Read existing results if file exists
            if self.output_file.exists():
                existing_df = pd.read_excel(self.output_file)
                existing_results = existing_df.to_dict('records')
            
            # Combine existing and new results
            all_results = existing_results + new_results
            
            # Write combined results
            self.write_results(all_results)
            
        except Exception as e:
            self.logger.error(f"Error appending results: {str(e)}")
            raise
    
    def validate_input_file(self, required_columns: List[str] = None) -> bool:
        """Validate input Excel file structure.
        
        Args:
            required_columns: List of required column names
            
        Returns:
            True if file is valid, False otherwise
        """
        if required_columns is None:
            required_columns = ["UEN"]
        
        try:
            if not self.input_file.exists():
                self.logger.error(f"Input file does not exist: {self.input_file}")
                return False
            
            df = pd.read_excel(self.input_file)
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                self.logger.error(f"Missing required columns: {missing_columns}")
                return False
            
            if df.empty:
                self.logger.error("Input file is empty")
                return False
            
            self.logger.info("Input file validation successful")
            return True
            
        except Exception as e:
            self.logger.error(f"Error validating input file: {str(e)}")
            return False
    
    def create_sample_input(self, sample_uens: List[str] = None):
        """Create a sample input Excel file.
        
        Args:
            sample_uens: List of sample UENs to include
        """
        if sample_uens is None:
            sample_uens = [
                "200012345A",
                "199812345B", 
                "202112345C"
            ]
        
        try:
            # Create sample DataFrame
            df = pd.DataFrame({"UEN": sample_uens})
            
            # Create input directory if it doesn't exist
            self.input_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Write sample file
            df.to_excel(self.input_file, index=False)
            
            self.logger.info(f"Created sample input file: {self.input_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating sample input file: {str(e)}")
            raise